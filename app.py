from __future__ import annotations

import copy
import io
import json
import math
import re
import time
import unicodedata
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.shapes import Drawing, Line, Rect, String
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from streamlit_local_storage import LocalStorage
except ImportError:  # Local fallback keeps the app usable without the optional browser-storage package.
    LocalStorage = None

# =============================================================================
# PAGE + DESIGN
# =============================================================================

st.set_page_config(
    page_title="Wall Assembly Embodied Carbon Calculator",
    layout="wide",
    initial_sidebar_state="expanded",
)

COLORS = {
    "blue": "#002F65",
    "green": "#C0D83F",
    "purple": "#86048A",
    "background": "#F5F7FA",
    "paper": "#FFFFFF",
    "ink": "#17253A",
    "muted": "#64748B",
    "line": "#D8E0E8",
    "soft_blue": "#EAF1F8",
    "soft_green": "#F2F8D5",
    "soft_purple": "#F5EAF6",
}

STARTING_CATEGORIES = [
    "Exterior Cladding",
    "Attachment System",
    "Moisture and Air Control",
    "Sheathing",
    "Insulation",
    "Stud and Framing",
    "Interior Finishes",
]

CATEGORY_DESCRIPTIONS = {
    "Exterior Cladding": "Exterior finish materials, including brick, panels, CMU, and concrete masonry.",
    "Attachment System": "Rails, clips, supports, reinforcing, ties, and other attachment components.",
    "Insulation": "Continuous, cavity, batt, board, and other thermal insulation.",
    "Stud and Framing": "Wood studs, steel studs, backup framing, and strapping used as framing.",
    "Moisture and Air Control": "Air barriers, vapor control layers, and weather-resistive membranes.",
    "Sheathing": "Plywood, gypsum sheathing, and other wall sheathing products.",
    "Interior Finishes": "Gypsum board, plaster, paint, and other interior finish layers.",
}

CATEGORY_COLORS = {
    "Exterior Cladding": "#002F65",
    "Attachment System": "#86048A",
    "Insulation": "#C0D83F",
    "Stud and Framing": "#2D5C8C",
    "Moisture and Air Control": "#6C8DB3",
    "Sheathing": "#A44CA8",
    "Interior Finishes": "#667085",
}

st.markdown(
    """
    <style>
      @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;500;600;700;800&display=swap');

      /* Apply Roboto through inheritance. Do not target every span/div because
         Streamlit uses a separate Material Symbols font for its interface icons. */
      html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stSidebar"] {
        font-family: 'Roboto', Arial, sans-serif;
      }
      button, input, textarea, label, select,
      h1, h2, h3, h4, h5, h6, p, li, td, th {
        font-family: 'Roboto', Arial, sans-serif;
      }

      /* Restore Streamlit's icon font. Without this, icon names such as
         arrow_right and upload appear as overlapping words. */
      [data-testid="stIconMaterial"],
      .material-symbols-rounded,
      .material-symbols-outlined,
      span[class*="material-symbols"] {
        font-family: 'Material Symbols Rounded', 'Material Symbols Outlined' !important;
        font-weight: normal !important;
        font-style: normal !important;
        letter-spacing: normal !important;
        text-transform: none !important;
        white-space: nowrap !important;
        word-wrap: normal !important;
        direction: ltr !important;
        -webkit-font-feature-settings: 'liga' !important;
        -webkit-font-smoothing: antialiased !important;
        font-feature-settings: 'liga' !important;
      }

      .stApp { background: #F5F7FA; color: #17253A; }
      [data-testid="stHeader"] { background: rgba(245, 247, 250, 0.92); }
      [data-testid="stSidebar"] {
        background: #002F65;
        border-right: 0;
      }

      /* Keep the collapse arrow visible while the purple sidebar is open. */
      [data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button,
      [data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button *,
      [data-testid="stSidebar"] button[kind="header"],
      [data-testid="stSidebar"] button[kind="header"] * {
        color: #FFFFFF !important;
      }
      [data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] svg,
      [data-testid="stSidebar"] button[kind="header"] svg {
        fill: #FFFFFF !important;
        stroke: #FFFFFF !important;
      }

      /* Sidebar headings and regular copy remain white. */
      [data-testid="stSidebar"] h1,
      [data-testid="stSidebar"] h2,
      [data-testid="stSidebar"] h3,
      [data-testid="stSidebar"] h4,
      [data-testid="stSidebar"] > div > div > div > p,
      [data-testid="stSidebar"] label,
      [data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
        color: #FFFFFF !important;
      }
      [data-testid="stSidebar"] input,
      [data-testid="stSidebar"] textarea {
        color: #17253A !important;
      }
      [data-testid="stSidebar"] div[data-baseweb="select"] * {
        color: #17253A;
      }
      .sidebar-brand {
        color: #FFFFFF;
        font-size: 1.08rem;
        font-weight: 800;
        line-height: 1.28;
        margin: 2px 0 18px;
      }
      [data-testid="stSidebar"] [data-testid="stRadio"] label,
      [data-testid="stSidebar"] [data-testid="stRadio"] label p,
      [data-testid="stSidebar"] [data-testid="stRadio"] label span:not([data-testid="stIconMaterial"]) {
        color: #FFFFFF !important;
        opacity: 1 !important;
      }
      [data-testid="stSidebar"] [data-testid="stRadio"] label p {
        font-weight: 500;
      }
      [data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.34) !important;
      }

      /* Sidebar buttons and uploader need dark text on white surfaces. */
      [data-testid="stSidebar"] .stButton > button,
      [data-testid="stSidebar"] .stDownloadButton > button {
        width: 100%;
        min-height: 42px;
        background: #FFFFFF !important;
        color: #86048A !important;
        border: 1px solid rgba(255,255,255,0.75) !important;
        border-radius: 10px !important;
        box-shadow: none !important;
      }
      [data-testid="stSidebar"] .stButton > button:hover,
      [data-testid="stSidebar"] .stDownloadButton > button:hover {
        background: #F1F5F9 !important;
        border-color: #C0D83F !important;
      }
      [data-testid="stSidebar"] .stButton > button p,
      [data-testid="stSidebar"] .stButton > button span:not([data-testid="stIconMaterial"]),
      [data-testid="stSidebar"] .stDownloadButton > button p,
      [data-testid="stSidebar"] .stDownloadButton > button span:not([data-testid="stIconMaterial"]) {
        color: #002F65 !important;
      }
      [data-testid="stSidebar"] button:disabled,
      [data-testid="stSidebar"] button:disabled p,
      [data-testid="stSidebar"] button:disabled span:not([data-testid="stIconMaterial"]) {
        color: #718096 !important;
        background: #E9EEF3 !important;
        opacity: 1 !important;
      }

      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
        background: #FFFFFF !important;
        border: 1px dashed #AFC1D5 !important;
        border-radius: 12px !important;
        padding: 14px !important;
      }
      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] p,
      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small,
      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] span:not([data-testid="stIconMaterial"]) {
        color: #53657A !important;
      }
      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
        background: #F5F7FA !important;
        color: #002F65 !important;
        border: 1px solid #CCD6E0 !important;
      }
      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button p,
      [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button span:not([data-testid="stIconMaterial"]) {
        color: #002F65 !important;
      }

      [data-testid="stMetric"] {
        background: #FFFFFF;
        border: 1px solid #D8E0E8;
        border-radius: 14px;
        padding: 14px 16px;
        box-shadow: 0 4px 14px rgba(0, 47, 101, 0.06);
      }
      [data-testid="stMetricLabel"] { color: #64748B; font-weight: 600; }
      [data-testid="stMetricValue"] { color: #002F65; font-weight: 800; }

      .hero {
        position: relative;
        overflow: hidden;
        background: linear-gradient(122deg, #002F65 0%, #002F65 64%, #86048A 130%);
        border-radius: 20px;
        padding: 28px 30px;
        margin-bottom: 20px;
        color: white;
        box-shadow: 0 12px 30px rgba(0, 47, 101, 0.18);
      }
      .hero::after {
        content: "";
        position: absolute;
        width: 180px;
        height: 180px;
        right: -62px;
        top: -75px;
        border-radius: 50%;
        background: #C0D83F;
        opacity: 0.92;
      }
      .hero h1 { position: relative; z-index: 1; font-size: 2.1rem; margin: 0; letter-spacing: -0.025em; }
      .hero p { position: relative; z-index: 1; margin: 8px 0 0; opacity: 0.88; max-width: 820px; }

      .section-label {
        color: #002F65;
        font-size: 0.76rem;
        font-weight: 800;
        letter-spacing: 0.11em;
        text-transform: uppercase;
        margin: 12px 0 7px;
      }
      .category-heading {
        display: flex;
        align-items: center;
        gap: 10px;
        margin-bottom: 3px;
      }
      .category-dot {
        width: 10px;
        height: 10px;
        border-radius: 999px;
        flex: 0 0 10px;
      }
      .category-title { color: #002F65; font-size: 1.04rem; font-weight: 800; }
      .category-copy { color: #64748B; font-size: 0.86rem; margin-bottom: 10px; }

      .epd-card {
        background: #FFFFFF;
        border: 1px solid #D8E0E8;
        border-left: 5px solid #C0D83F;
        border-radius: 12px;
        padding: 12px 14px;
        margin: 8px 0 12px;
      }
      .epd-card strong { color: #002F65; }
      .small-muted { color: #64748B; font-size: 0.86rem; }
      .formula-box {
        background: #EAF1F8;
        border: 1px solid #D5E2EF;
        border-radius: 10px;
        padding: 11px 13px;
        color: #002F65;
        font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
        font-size: 0.86rem;
      }
      .layer-empty {
        background: #F8FAFC;
        border: 1px dashed #C9D4E0;
        border-radius: 11px;
        padding: 11px 13px;
        color: #64748B;
        font-size: 0.88rem;
      }

      div[data-testid="stExpander"] {
        background: #FFFFFF;
        border: 1px solid #D8E0E8;
        border-radius: 14px;
        overflow: hidden;
        box-shadow: 0 3px 12px rgba(0, 47, 101, 0.035);
      }
      div[data-testid="stExpander"] summary {
        min-height: 48px;
        align-items: center;
      }
      div[data-testid="stExpander"] summary p {
        color: #17253A !important;
        line-height: 1.35;
        margin: 0;
      }
      div[data-testid="stVerticalBlockBorderWrapper"] {
        border-color: #D8E0E8 !important;
        border-radius: 16px !important;
        background: #FFFFFF;
      }

      /* Keep select text on one clean line without covering the chevron. */
      div[data-baseweb="select"] > div {
        min-height: 42px;
      }
      div[data-baseweb="select"] span:not([data-testid="stIconMaterial"]) {
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
      }

      .stButton > button, .stDownloadButton > button {
        border-radius: 10px;
        font-weight: 700;
        border-color: #002F65;
      }
      .stButton > button[kind="primary"] {
        background: #C0D83F;
        color: #002F65;
        border-color: #C0D83F;
      }

      /* Prevent data-grid menus from inheriting narrow text metrics. */
      [data-testid="stDataFrame"] {
        font-family: 'Roboto', Arial, sans-serif;
      }

      hr { border-color: #D8E0E8; }
    </style>
    """,
    unsafe_allow_html=True,
)

# =============================================================================
# MATERIAL LIBRARY — EXTRACTED FROM THE EXCEL WORKBOOK
# =============================================================================

PRESET_MATERIALS: list[dict[str, Any]] = [{'category': 'Exterior Cladding',
  'name': 'Clay Brick , Randers Tegl',
  'emitted': 8.491,
  'stored': 0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Wood-fired bricks, HG Matthews',
  'emitted': 15.61,
  'stored': -0.352,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Brick , Glen-Gery',
  'emitted': 47.6,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Clay Masonry Products Industry Average EPD',
  'emitted': 31.8,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'West Jordan Clay Masonry Products, Interstate Brick',
  'emitted': 52.1,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Hardie Plank, James Hardie',
  'emitted': 7.17,
  'stored': -1.85,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Hardie Panel, Architectural Panel, James Hardie',
  'emitted': 7.51,
  'stored': -1.85,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Swisspearl Carat/Avera/Vintago/Reflex',
  'emitted': 14.52,
  'stored': -0.9091,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Formica Compact PROTEC+ 4mm, Nemho',
  'emitted': 15.5,
  'stored': -7.29,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Trespa Meteon EDS grade - 8mm, Nemho',
  'emitted': 18.8,
  'stored': -15.1,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Meteon EDF grade 6mm, Trespa International',
  'emitted': 19.9,
  'stored': -9.13,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Fundermax Max Compact Exterior',
  'emitted': 24.0,
  'stored': -8.85,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Trespa Meteon HPL, Trespa International',
  'emitted': 26.5,
  'stored': -12.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Ancon Masonry Support Systems, Leviat Ltd',
  'emitted': 17.55,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Thin Tech Elite Cladding System, Glen-Gery',
  'emitted': 22.44,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'CFS Adjustable Masonry Support, Construction Fixing Systems Limited',
  'emitted': 84.6,
  'stored': -0.0729,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Reveal Support Plate, IG Masonry Support',
  'emitted': 5.71,
  'stored': 0.0,
  'unit': 'kgCO₂e/unit',
  'method': 'Unit Grid',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Welded Masonry Support System, IG Masonry Support',
  'emitted': 141,
  'stored': 0.0,
  'unit': 'kgCO₂e/unit',
  'method': 'Unit Grid',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Wood Strapping, Roseburg Forest Products',
  'emitted': 34.2,
  'stored': -727.2,
  'unit': 'kgCO₂e/m³',
  'method': 'Wood Stud Framing',
  'source': 'Verified against 2026_06_29 New Beam Results',
  'source_declared_unit': '1 m³ softwood lumber'},
 {'category': 'Attachment System',
  'name': 'NU-Wall Aluminum Cladding Ancillaries Anodized, Nu-Wall Cladding',
  'emitted': 2.311,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Green Girt, Advanced Architectural Products',
  'emitted': 4.208,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Cascadia Clip',
  'emitted': 6.463,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Existing calculator preset',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'KWS PanelRail and RevealRail, Knight Wall Systems',
  'emitted': 6.619,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'MFI Cladding Support System, Knight Wall',
  'emitted': 6.561679790026246,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Cladding Rails, Generic Primary Aluminum',
  'emitted': 17.424,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'TimberBoard Rigid Exterior Continuous Insulation, TimberHP',
  'emitted': 0.0,
  'stored': -8.377,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'PINK Next Gen Fiberglas Insulation, Owens Corning',
  'emitted': 0.6645,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'EcoBatt and EcoRoll Insulation, Knauf Insulation',
  'emitted': 0.997,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'K-13 and K-13 High-R System, International Cellulose Corporation',
  'emitted': 1.29,
  'stored': -9.354,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'Thermafiber Mineral Wool, Owens Corning',
  'emitted': 1.333,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'Expanded Polystyrene Insulation, EPS Industry Alliance',
  'emitted': 2.507,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'Polyisocyanurate Foam Trymer 2000, Johns Manville',
  'emitted': 2.89,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'PIMA Polyiso Wall Insulation',
  'emitted': 4.16,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'Foamular NGX XPS Insulation',
  'emitted': 6.58,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'JM CladStone 45 Water & Fire Block Insulation, Johns Manville',
  'emitted': 2.9,
  'stored': 0.0,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Insulation',
  'name': 'TimberBatt Wood Fiber Batt Insulation, TimberHP',
  'emitted': 0.0,
  'stored': -2.47,
  'unit': 'kgCO₂e/m² at R/RSI',
  'method': 'Insulation R-value',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Stud and Framing',
  'name': 'Wood Stud, North American Softwood Lumber',
  'emitted': 34.2,
  'stored': -727.2,
  'unit': 'kgCO₂e/m³',
  'method': 'Wood Stud Framing',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Stud and Framing',
  'name': 'Steel Stud, Cold-Formed Steel Framing Industry Average',
  'emitted': 2.44,
  'stored': 0.0,
  'unit': 'kgCO₂e/kg',
  'method': 'Steel Stud Framing',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Moisture and Air Control',
  'name': 'Continuous Self-Adhered Vapor Open Air Barrier Membrane, SOPREMA / SIGA',
  'emitted': 1.54,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Sheathing',
  'name': 'Plywood, Industry Average',
  'emitted': 219.32,
  'stored': -862.9,
  'unit': 'kgCO₂e/m³',
  'method': 'Volume Layer',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Sheathing',
  'name': '5/8 in DensGlass Fireguard Sheathing Gypsum Panel',
  'emitted': 5.4,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'CMU, Genest Concrete Works Inc.',
  'emitted': 129.0,
  'stored': 0.0,
  'unit': 'kgCO₂e/m³',
  'method': 'Volume Layer',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'CMU: Medium Weight Concrete Masonry Units Sun Valley Plant, Angelus Block Co',
  'emitted': 157.0,
  'stored': 0.0,
  'unit': 'kgCO₂e/m³',
  'method': 'Volume Layer',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Normal Weight and Lightweight Concrete Block Masonry Unit',
  'emitted': 176.54,
  'stored': 0.0,
  'unit': 'kgCO₂e/m³',
  'method': 'Volume Layer',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'CMU: NWCX, Gardner Plant, Jandris Block',
  'emitted': 180.0,
  'stored': 0.0,
  'unit': 'kgCO₂e/m³',
  'method': 'Volume Layer',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Exterior Cladding',
  'name': 'Concrete Masonry Unit (CMU), hollow-core, normal-weight industry average',
  'emitted': 208.0,
  'stored': 0.0,
  'unit': 'kgCO₂e/m³',
  'method': 'Volume Layer',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Attachment System',
  'name': 'Steel Wall Ties, Arminox A/S',
  'emitted': 0.3465,
  'stored': 0.0,
  'unit': 'kgCO₂e/m',
  'method': 'Linear Members',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Other',
  'name': 'Gold Bond / Industry-Average 5/8 in Type X Gypsum Board',
  'emitted': 2.98,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Other',
  'name': 'Gypsum Plaster Finish / Kal-Kore Substitute',
  'emitted': 0.12,
  'stored': 0.0,
  'unit': 'kgCO₂e/kg',
  'method': 'Mass by Grammage',
  'source': 'Preset material library',
  'source_declared_unit': ''},
 {'category': 'Other',
  'name': 'Acrylic Latex Paint — generic substitute',
  'emitted': 2.19,
  'stored': 0.0,
  'unit': 'kgCO₂e/m²',
  'method': 'Area',
  'source': 'Preset material library',
  'source_declared_unit': ''}]

DECLARED_UNITS = [
    "kgCO₂e/m²",
    "kgCO₂e/m³",
    "kgCO₂e/kg",
    "kgCO₂e/m",
    "kgCO₂e/unit",
    "kgCO₂e/m² at R/RSI",
]

AUTO_METHOD_BY_UNIT = {
    "kgCO₂e/m²": "Area",
    "kgCO₂e/m³": "Volume Layer",
    "kgCO₂e/kg": "Mass by Grammage",
    "kgCO₂e/m": "Linear Members",
    "kgCO₂e/unit": "Unit Grid",
    "kgCO₂e/m² at R/RSI": "Insulation R-value",
}


def method_from_declared_unit(unit: str | None) -> str:
    """Return the standard calculation method implied by an EPD declared unit."""
    return AUTO_METHOD_BY_UNIT.get(unit or "kgCO₂e/m²", "Area")


WOOD_SIZES = {
    "1x2": (1, 2),
    "1x3": (1, 3),
    "1x4": (1, 4),
    "2x2": (2, 2),
    "2x3": (2, 3),
    "2x4": (2, 5),
    "2x6": (2, 6),
    "2x8": (2, 8),
    "2x10": (2, 10),
    "2x12": (2, 12),
}

STEEL_GAUGE_MM = {
    "25": 0.455,
    "22": 0.752,
    "20": 0.912,
    "18": 1.090,
    "16": 1.370,
    "14": 1.720,
}

STEEL_DENSITY_KG_M3 = 7850.0
STEEL_FLANGE_IN = 1.625
STEEL_LIP_IN = 0.5

FT2_TO_M2 = 0.09290304
IN_TO_M = 0.0254
FT_TO_M = 0.3048
FT3_TO_M3 = 0.028316846592
R_TO_RSI = 0.1761101838

# =============================================================================
# DATA MODEL + SESSION STATE
# =============================================================================

CATEGORY_ALIASES = {
    "Masonry and Concrete": "Exterior Cladding",
    "Reinforcing and Ties": "Attachment System",
    "Stud and Backup Framing": "Stud and Framing",
    "Finish": "Interior Finishes",
    "Other": "Interior Finishes",
}

CUSTOM_PRESET_NAMES = {"Custom EPD", "Other (custom EPD)"}
EMPTY_PRESET_NAMES = {"", "None", None}


def normalized_category(category: str | None) -> str:
    value = CATEGORY_ALIASES.get(category or "Interior Finishes", category or "Interior Finishes")
    return value if value in STARTING_CATEGORIES else "Interior Finishes"


# Normalize legacy preset categories once at startup.
for _preset_item in PRESET_MATERIALS:
    _preset_item["category"] = normalized_category(_preset_item.get("category"))


def new_material(category: str, is_base: bool = False) -> dict[str, Any]:
    return {
        "id": uuid.uuid4().hex,
        "category": normalized_category(category),
        "is_base": is_base,
        "preset": "None",
        "custom_name": "",
        "custom_emitted": 0.0,
        "custom_stored": 0.0,
        "custom_unit": "kgCO₂e/m²",
        "inputs": {},
    }


def new_assembly(number: int = 1) -> dict[str, Any]:
    return {
        "id": uuid.uuid4().hex,
        "name": f"Wall Option {number}",
        "width_ft": 10.0,
        "height_ft": 10.0,
        "materials": [new_material(category, is_base=True) for category in STARTING_CATEGORIES],
    }


def normalize_material(material: dict[str, Any]) -> dict[str, Any]:
    material.setdefault("id", uuid.uuid4().hex)
    material["category"] = normalized_category(material.get("category"))
    preset = material.get("preset")
    if preset in EMPTY_PRESET_NAMES:
        material["preset"] = "None"
    elif preset == "Other (custom EPD)":
        material["preset"] = "Custom EPD"
    material.setdefault("is_base", False)
    material.setdefault("custom_name", "")
    material.setdefault("custom_emitted", 0.0)
    material.setdefault("custom_stored", 0.0)
    material.setdefault("custom_unit", "kgCO₂e/m²")
    material.setdefault("inputs", {})
    return material


def normalize_assembly(assembly: dict[str, Any]) -> dict[str, Any]:
    assembly.setdefault("id", uuid.uuid4().hex)
    assembly.setdefault("name", "Wall Option")
    assembly.setdefault("width_ft", 10.0)
    assembly.setdefault("height_ft", 10.0)
    materials = [normalize_material(item) for item in assembly.get("materials", [])]

    for category in STARTING_CATEGORIES:
        matches = [item for item in materials if item["category"] == category]
        if matches:
            matches[0]["is_base"] = True
        else:
            materials.append(new_material(category, is_base=True))

    category_order = {category: index for index, category in enumerate(STARTING_CATEGORIES)}
    materials.sort(key=lambda item: (category_order[item["category"]], not item.get("is_base", False)))
    assembly["materials"] = materials
    return assembly


def normalize_custom_library(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for item in items:
        copied = dict(item)
        copied.setdefault("id", uuid.uuid4().hex)
        copied["category"] = normalized_category(copied.get("category"))
        copied.setdefault("name", "Custom material")
        copied.setdefault("emitted", 0.0)
        copied.setdefault("stored", 0.0)
        copied.setdefault("unit", "kgCO₂e/m²")
        copied["method"] = method_from_declared_unit(copied["unit"])
        copied.setdefault("source", "Custom material")
        copied.setdefault("source_declared_unit", "")
        normalized.append(copied)
    return normalized


CUSTOM_LIBRARY_STORAGE_KEY = "wall_embodied_carbon_custom_materials_v1"
CUSTOM_LIBRARY_FALLBACK_FILE = Path(__file__).with_name(".wall_custom_materials.json")

try:
    BROWSER_STORAGE = LocalStorage() if LocalStorage is not None else None
except Exception:
    BROWSER_STORAGE = None


def _decode_custom_library(value: Any) -> list[dict[str, Any]] | None:
    if value in (None, "", []):
        return None
    try:
        if isinstance(value, str):
            value = json.loads(value)
        if isinstance(value, dict):
            value = value.get("value", value.get("data", value.get("custom_library")))
            if isinstance(value, str):
                value = json.loads(value)
        if isinstance(value, list):
            return normalize_custom_library(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    return None


def _load_fallback_custom_library() -> list[dict[str, Any]]:
    try:
        if CUSTOM_LIBRARY_FALLBACK_FILE.exists():
            data = json.loads(CUSTOM_LIBRARY_FALLBACK_FILE.read_text(encoding="utf-8"))
            return normalize_custom_library(data if isinstance(data, list) else [])
    except (OSError, ValueError, json.JSONDecodeError):
        pass
    return []


def persist_custom_library() -> None:
    st.session_state.custom_library = normalize_custom_library(st.session_state.custom_library)
    payload = json.dumps(st.session_state.custom_library, ensure_ascii=False)
    if BROWSER_STORAGE is not None:
        st.session_state._custom_storage_write_number = st.session_state.get("_custom_storage_write_number", 0) + 1
        try:
            BROWSER_STORAGE.setItem(
                CUSTOM_LIBRARY_STORAGE_KEY,
                payload,
                key=f"custom_library_write_{st.session_state._custom_storage_write_number}",
            )
        except TypeError:
            BROWSER_STORAGE.setItem(CUSTOM_LIBRARY_STORAGE_KEY, payload)
        # The browser component needs a moment to write before a forced rerun.
        time.sleep(0.8)
    else:
        try:
            CUSTOM_LIBRARY_FALLBACK_FILE.write_text(payload, encoding="utf-8")
        except OSError:
            pass


def hydrate_custom_library() -> None:
    if st.session_state.get("_custom_library_hydrated"):
        return

    if BROWSER_STORAGE is not None:
        try:
            returned = BROWSER_STORAGE.getItem(
                CUSTOM_LIBRARY_STORAGE_KEY,
                key="custom_library_storage_read",
            )
        except TypeError:
            returned = BROWSER_STORAGE.getItem(CUSTOM_LIBRARY_STORAGE_KEY)
        stored = st.session_state.get("custom_library_storage_read", returned)
        decoded = _decode_custom_library(stored)
        if decoded is not None and not st.session_state.get("_custom_library_dirty"):
            st.session_state.custom_library = decoded
            st.session_state._custom_library_hydrated = True
    else:
        st.session_state.custom_library = _load_fallback_custom_library()
        st.session_state._custom_library_hydrated = True


def mark_custom_library_changed() -> None:
    st.session_state._custom_library_dirty = True
    st.session_state._custom_library_hydrated = True
    persist_custom_library()


def update_custom_material_references(old_name: str, new_name: str | None) -> None:
    replacement = new_name or "None"
    for assembly in st.session_state.assemblies:
        for material in assembly.get("materials", []):
            if material.get("preset") == old_name:
                material["preset"] = replacement


def initialize_state() -> None:
    if "assemblies" not in st.session_state:
        first = new_assembly(1)
        st.session_state.assemblies = [first]
        st.session_state.active_assembly_id = first["id"]
    else:
        st.session_state.assemblies = [normalize_assembly(item) for item in st.session_state.assemblies]

    if "custom_library" not in st.session_state:
        st.session_state.custom_library = []
    else:
        st.session_state.custom_library = normalize_custom_library(st.session_state.custom_library)

    if "page" not in st.session_state:
        st.session_state.page = "Build Assemblies"
    if "show_add_custom" not in st.session_state:
        st.session_state.show_add_custom = False
    if "editing_custom_id" not in st.session_state:
        st.session_state.editing_custom_id = None


initialize_state()
hydrate_custom_library()


def all_library_materials() -> list[dict[str, Any]]:
    return PRESET_MATERIALS + st.session_state.custom_library


def material_lookup(name: str) -> dict[str, Any] | None:
    for item in all_library_materials():
        if item["name"] == name:
            return item
    return None


def material_name_exists(name: str, exclude_custom_id: str | None = None) -> bool:
    target = name.strip().casefold()
    if not target:
        return False
    for item in PRESET_MATERIALS:
        if str(item.get("name", "")).strip().casefold() == target:
            return True
    for item in st.session_state.custom_library:
        if item.get("id") == exclude_custom_id:
            continue
        if str(item.get("name", "")).strip().casefold() == target:
            return True
    return False


def category_list() -> list[str]:
    return STARTING_CATEGORIES.copy()


def presets_for_category(category: str) -> list[str]:
    names = [item["name"] for item in all_library_materials() if item["category"] == category]
    return ["None"] + names + ["Custom EPD"]


def material_is_selected(material: dict[str, Any]) -> bool:
    return material.get("preset") not in EMPTY_PRESET_NAMES


def get_active_assembly() -> dict[str, Any]:
    active = st.session_state.active_assembly_id
    for assembly in st.session_state.assemblies:
        if assembly["id"] == active:
            return assembly
    st.session_state.active_assembly_id = st.session_state.assemblies[0]["id"]
    return st.session_state.assemblies[0]


def select_index(options: list[Any], value: Any, fallback: int = 0) -> int:
    try:
        return options.index(value)
    except ValueError:
        return fallback


def input_value(material: dict[str, Any], key: str, default: Any) -> Any:
    return material.setdefault("inputs", {}).get(key, default)


def set_input(material: dict[str, Any], key: str, value: Any) -> Any:
    material.setdefault("inputs", {})[key] = value
    return value

# =============================================================================
# CALCULATION ENGINE — DIRECT TRANSLATION OF THE EXCEL FORMULAS
# =============================================================================


def member_count(span_ft: float, spacing_in: float, include_edges: str) -> int:
    if spacing_in <= 0:
        raise ValueError("Spacing must be greater than zero.")
    return max(1, math.ceil(span_ft * 12.0 / spacing_in) + (1 if include_edges == "Yes" else 0))


def effective_epd(material: dict[str, Any]) -> dict[str, Any] | None:
    if material.get("preset") in CUSTOM_PRESET_NAMES:
        return {
            "category": material.get("category", "Interior Finishes"),
            "name": material.get("custom_name", "").strip() or "Custom material",
            "emitted": float(material.get("custom_emitted", 0.0)),
            "stored": float(material.get("custom_stored", 0.0)),
            "unit": material.get("custom_unit", "kgCO₂e/m²"),
            "method": method_from_declared_unit(material.get("custom_unit", "kgCO₂e/m²")),
            "source": "Custom EPD entered by user",
            "source_declared_unit": "",
        }
    return material_lookup(material.get("preset", ""))


def calculate_material(assembly: dict[str, Any], material: dict[str, Any]) -> dict[str, Any]:
    width_ft = float(assembly.get("width_ft", 0.0))
    height_ft = float(assembly.get("height_ft", 0.0))
    area_ft2 = width_ft * height_ft
    area_m2 = area_ft2 * FT2_TO_M2

    base = {
        "valid": False,
        "status": "Not selected" if not material_is_selected(material) else "Check inputs",
        "error": "",
        "quantity": 0.0,
        "quantity_unit": "",
        "emitted": 0.0,
        "stored": 0.0,
        "net": 0.0,
        "emitted_intensity": 0.0,
        "net_intensity": 0.0,
        "name": material.get("custom_name") or material.get("preset") or "Unselected material",
        "category": material.get("category", "Interior Finishes"),
        "method": "",
        "unit": "",
        "formula_note": "",
    }

    if not material_is_selected(material):
        base["name"] = "None"
        return base
    if width_ft <= 0 or height_ft <= 0:
        base["error"] = "Enter wall width and height."
        return base

    epd = effective_epd(material)
    if epd is None:
        base["status"] = "Select material"
        base["error"] = "Select a preset material or choose Custom EPD."
        return base
    if material.get("preset") in CUSTOM_PRESET_NAMES and not material.get("custom_name", "").strip():
        base["status"] = "Complete custom EPD"
        base["error"] = "Enter a custom material name."
        return base

    method = epd["method"]
    inputs = material.get("inputs", {})
    quantity = 0.0
    quantity_unit = ""
    note = ""

    try:
        if method == "Area":
            quantity = area_m2
            quantity_unit = "m²"
            note = "wall area (ft²) × 0.09290304"

        elif method == "Volume Layer":
            thickness_in = float(inputs.get("thickness_in", 0.0))
            if thickness_in <= 0:
                raise ValueError("Enter product thickness in inches.")
            quantity = area_m2 * thickness_in * IN_TO_M
            quantity_unit = "m³"
            note = "wall area (m²) × thickness (in) × 0.0254"

        elif method == "Mass by Grammage":
            mass_input_method = inputs.get("mass_input_method", "Use grammage (kg/m²)")
            if mass_input_method == "Enter total mass (kg)":
                quantity = float(inputs.get("total_mass_kg", 0.0))
                if quantity <= 0:
                    raise ValueError("Enter the total installed mass in kg.")
                note = "total installed mass (kg)"
            else:
                grammage = float(inputs.get("grammage_kg_m2", 0.0))
                if grammage <= 0:
                    raise ValueError("Enter material grammage in kg/m².")
                quantity = area_m2 * grammage
                note = "wall area (m²) × grammage (kg/m²)"
            quantity_unit = "kg"

        elif method == "Mass by Volume + Density":
            thickness_in = float(inputs.get("thickness_in", 0.0))
            density = float(inputs.get("density_kg_m3", 0.0))
            if thickness_in <= 0 or density <= 0:
                raise ValueError("Enter thickness and density.")
            quantity = area_m2 * thickness_in * IN_TO_M * density
            quantity_unit = "kg"
            note = "area (m²) × thickness (m) × density (kg/m³)"

        elif method == "Mass by Item Grid":
            horizontal = float(inputs.get("horizontal_spacing_in", 0.0))
            vertical = float(inputs.get("vertical_spacing_in", 0.0))
            mass_each = float(inputs.get("mass_per_item_kg", 0.0))
            include_edges = inputs.get("include_wall_edges", "Yes")
            if horizontal <= 0 or vertical <= 0 or mass_each <= 0:
                raise ValueError("Enter horizontal spacing, vertical spacing, and mass per item.")
            columns = member_count(width_ft, horizontal, include_edges)
            rows = member_count(height_ft, vertical, include_edges)
            quantity = columns * rows * mass_each
            quantity_unit = "kg"
            note = "grid columns × grid rows × mass per item"

        elif method == "Mass by Linear Weight":
            orientation = inputs.get("orientation", "Vertical")
            spacing = float(inputs.get("spacing_in", 0.0))
            linear_weight = float(inputs.get("linear_weight_kg_m", 0.0))
            include_edges = inputs.get("include_edge_members", "Yes")
            if spacing <= 0 or linear_weight <= 0:
                raise ValueError("Enter spacing and linear weight.")
            span = width_ft if orientation == "Vertical" else height_ft
            length_each_ft = height_ft if orientation == "Vertical" else width_ft
            count = member_count(span, spacing, include_edges)
            quantity = count * length_each_ft * FT_TO_M * linear_weight
            quantity_unit = "kg"
            note = "member count × member length (m) × linear weight (kg/m)"

        elif method == "Linear Members":
            orientation = inputs.get("orientation", "Vertical")
            spacing = float(inputs.get("spacing_in", 0.0))
            include_edges = inputs.get("include_edge_members", "Yes")
            if spacing <= 0:
                raise ValueError("Enter member spacing.")
            span = width_ft if orientation == "Vertical" else height_ft
            length_each_ft = height_ft if orientation == "Vertical" else width_ft
            count = member_count(span, spacing, include_edges)
            quantity = count * length_each_ft * FT_TO_M
            quantity_unit = "m"
            note = "member count × member length (m)"

        elif method == "Unit Grid":
            use_horizontal = inputs.get("use_horizontal_spacing", "Yes") == "Yes"
            use_vertical = inputs.get("use_vertical_spacing", "Yes") == "Yes"
            include_edges = inputs.get("include_wall_edges", "Yes")

            columns = 1
            rows = 1
            directions = []

            if use_horizontal:
                horizontal = float(inputs.get("horizontal_spacing_in", 0.0))
                if horizontal <= 0:
                    raise ValueError("Enter horizontal spacing.")
                columns = member_count(width_ft, horizontal, include_edges)
                directions.append("horizontal count")

            if use_vertical:
                vertical = float(inputs.get("vertical_spacing_in", 0.0))
                if vertical <= 0:
                    raise ValueError("Enter vertical spacing.")
                rows = member_count(height_ft, vertical, include_edges)
                directions.append("vertical count")

            quantity = columns * rows
            quantity_unit = "units"
            if use_horizontal and use_vertical:
                note = "horizontal count × vertical count"
            elif directions:
                note = directions[0]
            else:
                note = "one unit for the wall"

        elif method == "Insulation R-value":
            required_r = float(inputs.get("required_r", 0.0))
            reference_type = inputs.get("epd_reference_type", "RSI (SI)")
            reference_value = float(inputs.get("epd_reference_value", 0.0))
            if required_r <= 0 or reference_value <= 0:
                raise ValueError("Enter required R-value and EPD reference R/RSI.")
            required_rsi = required_r * R_TO_RSI
            reference_rsi = reference_value * R_TO_RSI if reference_type == "R (US)" else reference_value
            quantity = area_m2 * required_rsi / reference_rsi
            quantity_unit = "m²-equivalent"
            note = "area (m²) × required RSI ÷ EPD reference RSI"

        elif method == "Wood Stud Framing":
            stud_size = inputs.get("stud_size", "2x4")
            spacing = float(inputs.get("spacing_in", 0.0))
            include_edges = inputs.get("include_edge_studs", "Yes")
            if stud_size not in WOOD_SIZES:
                raise ValueError("Choose a valid stud size.")
            if spacing <= 0:
                raise ValueError("Enter stud spacing.")
            count = member_count(width_ft, spacing, include_edges)
            actual_width_in, actual_depth_in = WOOD_SIZES[stud_size]
            cross_section_ft2 = actual_width_in * actual_depth_in / 144.0
            volume_ft3 = count * height_ft * cross_section_ft2
            quantity = volume_ft3 * FT3_TO_M3
            quantity_unit = "m³"
            note = "stud count × height × actual cross-section × ft³-to-m³"

        elif method == "Steel Stud Framing":
            depth_in = float(inputs.get("stud_depth_in", 0.0))
            gauge = str(inputs.get("gauge", "18"))
            spacing = float(inputs.get("spacing_in", 0.0))
            include_edges = inputs.get("include_edge_studs", "Yes")
            if depth_in <= 0 or spacing <= 0:
                raise ValueError("Enter stud depth and spacing.")
            if gauge not in STEEL_GAUGE_MM:
                raise ValueError("Choose a valid steel gauge.")
            count = member_count(width_ft, spacing, include_edges)
            total_length_m = count * height_ft * FT_TO_M
            strip_width_m = (depth_in + 2 * STEEL_FLANGE_IN + 2 * STEEL_LIP_IN) * IN_TO_M
            thickness_m = STEEL_GAUGE_MM[gauge] / 1000.0
            kg_per_m = strip_width_m * thickness_m * STEEL_DENSITY_KG_M3
            quantity = total_length_m * kg_per_m
            quantity_unit = "kg"
            note = "stud length × approximate C-section steel area × steel density"

        else:
            raise ValueError(f"Unsupported calculation method: {method}")

        emitted = quantity * float(epd["emitted"])
        stored = quantity * float(epd.get("stored", 0.0))
        net = emitted + stored
        emitted_intensity = emitted / area_m2 if area_m2 else 0.0
        net_intensity = net / area_m2 if area_m2 else 0.0

        base.update({
            "valid": True,
            "status": "Ready",
            "quantity": quantity,
            "quantity_unit": quantity_unit,
            "emitted": emitted,
            "stored": stored,
            "net": net,
            "emitted_intensity": emitted_intensity,
            "net_intensity": net_intensity,
            "name": epd["name"],
            "category": epd["category"],
            "method": method,
            "unit": epd["unit"],
            "formula_note": note,
            "epd_emitted": float(epd["emitted"]),
            "epd_stored": float(epd.get("stored", 0.0)),
            "source": epd.get("source", ""),
            "source_declared_unit": epd.get("source_declared_unit", ""),
        })
        return base

    except (TypeError, ValueError, ZeroDivisionError) as exc:
        base.update({
            "status": "Check inputs",
            "error": str(exc),
            "name": epd["name"],
            "category": epd["category"],
            "method": method,
            "unit": epd["unit"],
            "source": epd.get("source", ""),
        })
        return base


def calculate_assembly(assembly: dict[str, Any]) -> dict[str, Any]:
    rows = [calculate_material(assembly, material) for material in assembly.get("materials", [])]
    included = [row for row in rows if row["status"] != "Not selected"]
    ready = [row for row in rows if row["valid"]]
    width = float(assembly.get("width_ft", 0.0))
    height = float(assembly.get("height_ft", 0.0))
    area_ft2 = width * height
    area_m2 = area_ft2 * FT2_TO_M2
    emitted = sum(row["emitted"] for row in ready)
    stored = sum(row["stored"] for row in ready)
    net = emitted + stored
    return {
        "rows": rows,
        "area_ft2": area_ft2,
        "area_m2": area_m2,
        "emitted": emitted,
        "stored": stored,
        "net": net,
        "emitted_intensity": emitted / area_m2 if area_m2 else 0.0,
        "net_intensity": net / area_m2 if area_m2 else 0.0,
        "selected_count": len(included),
        "ready_count": len(ready),
        "status": "Not started" if not included else ("Ready" if len(included) == len(ready) else "Check inputs"),
    }

# =============================================================================
# INPUT RENDERING
# =============================================================================


def number_field(material: dict[str, Any], key: str, label: str, widget_key: str, default: float, step: float = 1.0, help_text: str | None = None) -> float:
    value = float(input_value(material, key, default))
    entered = st.number_input(label, min_value=0.0, value=value, step=step, key=widget_key, help=help_text)
    return set_input(material, key, float(entered))


def integer_field(material: dict[str, Any], key: str, label: str, widget_key: str, default: int = 0) -> int:
    value = int(input_value(material, key, default))
    entered = st.number_input(label, min_value=0, value=value, step=1, key=widget_key)
    return set_input(material, key, int(entered))


def dropdown_field(material: dict[str, Any], key: str, label: str, options: list[str], widget_key: str, default: str) -> str:
    current = input_value(material, key, default)
    entered = st.selectbox(label, options, index=select_index(options, current), key=widget_key)
    return set_input(material, key, entered)


def render_method_inputs(material: dict[str, Any], method: str, prefix: str) -> None:
    st.markdown('<div class="section-label">Required installation inputs</div>', unsafe_allow_html=True)

    if method == "Area":
        st.info("No additional input is required. The full wall area is used.")

    elif method == "Volume Layer":
        number_field(material, "thickness_in", "Product thickness (in)", f"{prefix}_thickness", 0.5, 0.125)

    elif method == "Mass by Grammage":
        mass_input_method = dropdown_field(
            material,
            "mass_input_method",
            "How do you know the installed mass?",
            ["Use grammage (kg/m²)", "Enter total mass (kg)"],
            f"{prefix}_mass_input_method",
            "Use grammage (kg/m²)",
        )
        if mass_input_method == "Enter total mass (kg)":
            number_field(
                material,
                "total_mass_kg",
                "Total installed mass (kg)",
                f"{prefix}_total_mass",
                1.0,
                0.1,
            )
            st.caption("Carbon = EPD GWP (kgCO₂e/kg) × total mass (kg).")
        else:
            number_field(
                material,
                "grammage_kg_m2",
                "Grammage (kg/m²)",
                f"{prefix}_grammage",
                1.0,
                0.1,
            )
            st.caption("Mass = wall area (m²) × grammage (kg/m²). Carbon = EPD GWP × mass.")

    elif method == "Mass by Volume + Density":
        c1, c2 = st.columns(2)
        with c1:
            number_field(material, "thickness_in", "Product thickness (in)", f"{prefix}_thickness", 0.5, 0.125)
        with c2:
            number_field(material, "density_kg_m3", "Density (kg/m³)", f"{prefix}_density", 500.0, 10.0)

    elif method == "Mass by Item Grid":
        c1, c2 = st.columns(2)
        with c1:
            number_field(material, "horizontal_spacing_in", "Horizontal spacing (in)", f"{prefix}_h_spacing", 16.0, 1.0)
        with c2:
            number_field(material, "vertical_spacing_in", "Vertical spacing (in)", f"{prefix}_v_spacing", 24.0, 1.0)
        c1, c2 = st.columns(2)
        with c1:
            number_field(material, "mass_per_item_kg", "Mass per item (kg)", f"{prefix}_mass_each", 0.25, 0.05)
        with c2:
            dropdown_field(material, "include_wall_edges", "Include wall edges?", ["Yes", "No"], f"{prefix}_wall_edges", "Yes")

    elif method == "Mass by Linear Weight":
        c1, c2 = st.columns(2)
        with c1:
            dropdown_field(material, "orientation", "Orientation", ["Vertical", "Horizontal"], f"{prefix}_orientation", "Vertical")
        with c2:
            number_field(material, "spacing_in", "Spacing (in)", f"{prefix}_spacing", 16.0, 1.0)
        c1, c2 = st.columns(2)
        with c1:
            number_field(material, "linear_weight_kg_m", "Linear weight (kg/m)", f"{prefix}_linear_weight", 1.0, 0.1)
        with c2:
            dropdown_field(material, "include_edge_members", "Include edge members?", ["Yes", "No"], f"{prefix}_edges", "Yes")

    elif method == "Linear Members":
        c1, c2 = st.columns(2)
        with c1:
            dropdown_field(material, "orientation", "Orientation", ["Vertical", "Horizontal"], f"{prefix}_orientation", "Vertical")
        with c2:
            number_field(material, "spacing_in", "Spacing (in)", f"{prefix}_spacing", 16.0, 1.0)
        dropdown_field(material, "include_edge_members", "Include edge members?", ["Yes", "No"], f"{prefix}_edges", "Yes")

    elif method == "Unit Grid":
        c1, c2 = st.columns(2)
        with c1:
            use_horizontal = dropdown_field(
                material,
                "use_horizontal_spacing",
                "Use horizontal spacing?",
                ["Yes", "No"],
                f"{prefix}_use_h_spacing",
                "Yes",
            )
        with c2:
            use_vertical = dropdown_field(
                material,
                "use_vertical_spacing",
                "Use vertical spacing?",
                ["Yes", "No"],
                f"{prefix}_use_v_spacing",
                "Yes",
            )

        spacing_columns = st.columns(2)
        if use_horizontal == "Yes":
            with spacing_columns[0]:
                number_field(material, "horizontal_spacing_in", "Horizontal spacing (in)", f"{prefix}_h_spacing", 16.0, 1.0)
        if use_vertical == "Yes":
            with spacing_columns[1]:
                number_field(material, "vertical_spacing_in", "Vertical spacing (in)", f"{prefix}_v_spacing", 24.0, 1.0)

        if use_horizontal == "Yes" or use_vertical == "Yes":
            dropdown_field(material, "include_wall_edges", "Include wall edges?", ["Yes", "No"], f"{prefix}_wall_edges", "Yes")
        else:
            st.info("One unit will be counted for the wall because neither spacing direction is used.")

    elif method == "Insulation R-value":
        c1, c2, c3 = st.columns(3)
        with c1:
            number_field(material, "required_r", "Required R-value", f"{prefix}_required_r", 20.0, 1.0)
        with c2:
            dropdown_field(material, "epd_reference_type", "EPD reference type", ["R (US)", "RSI (SI)"], f"{prefix}_reference_type", "RSI (SI)")
        with c3:
            number_field(material, "epd_reference_value", "EPD reference R / RSI", f"{prefix}_reference_value", 1.0, 0.1)

    elif method == "Wood Stud Framing":
        c1, c2, c3 = st.columns(3)
        with c1:
            dropdown_field(material, "stud_size", "Stud size", list(WOOD_SIZES), f"{prefix}_stud_size", "2x4")
        with c2:
            number_field(material, "spacing_in", "Stud spacing (in)", f"{prefix}_spacing", 16.0, 1.0)
        with c3:
            dropdown_field(material, "include_edge_studs", "Include edge studs?", ["Yes", "No"], f"{prefix}_edge_studs", "Yes")

    elif method == "Steel Stud Framing":
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            number_field(material, "stud_depth_in", "Stud depth (in)", f"{prefix}_depth", 6.0, 0.5)
        with c2:
            dropdown_field(material, "gauge", "Gauge", list(STEEL_GAUGE_MM), f"{prefix}_gauge", "18")
        with c3:
            number_field(material, "spacing_in", "Stud spacing (in)", f"{prefix}_spacing", 16.0, 1.0)
        with c4:
            dropdown_field(material, "include_edge_studs", "Include edge studs?", ["Yes", "No"], f"{prefix}_edge_studs", "Yes")
        st.caption("Steel mass uses an editable industry-average C-section approximation: 7,850 kg/m³ density, 1.625 in flanges, and 0.5 in lips.")

# =============================================================================
# EXPORTS
# =============================================================================



def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", name).strip() or "Wall Option"
    cleaned = cleaned[:31]
    original = cleaned
    counter = 2
    while cleaned in used:
        suffix = f" {counter}"
        cleaned = (original[:31 - len(suffix)] + suffix)
        counter += 1
    used.add(cleaned)
    return cleaned


def style_header(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="002F65")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDE2E4")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def workbook_bytes() -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Assembly Comparison"
    headers = [
        "Assembly", "Width (ft)", "Height (ft)", "Area (ft²)", "Materials Ready",
        "Emitted kgCO₂e", "Stored kgCO₂e", "Net kgCO₂e",
        "Emitted kgCO₂e/m²", "Net kgCO₂e/m²", "Status",
    ]
    summary.append(headers)
    style_header(summary, 1, 1, len(headers))

    used_names = {"Assembly Comparison"}
    for assembly in st.session_state.assemblies:
        result = calculate_assembly(assembly)
        summary.append([
            assembly["name"], assembly["width_ft"], assembly["height_ft"], result["area_ft2"],
            f'{result["ready_count"]}/{result["selected_count"]}', result["emitted"], result["stored"],
            result["net"], result["emitted_intensity"], result["net_intensity"], result["status"],
        ])

        ws = wb.create_sheet(safe_sheet_name(assembly["name"], used_names))
        ws["A1"] = "Wall Assembly Embodied Carbon"
        ws["A1"].font = Font(size=16, bold=True, color="002F65")
        ws["A3"], ws["B3"] = "Assembly", assembly["name"]
        ws["A4"], ws["B4"] = "Width (ft)", assembly["width_ft"]
        ws["A5"], ws["B5"] = "Height (ft)", assembly["height_ft"]
        ws["A6"], ws["B6"] = "Area (ft²)", result["area_ft2"]
        ws["D3"], ws["E3"] = "Total emitted", result["emitted"]
        ws["D4"], ws["E4"] = "Total stored", result["stored"]
        ws["D5"], ws["E5"] = "Net carbon", result["net"]
        ws["D6"], ws["E6"] = "Net intensity (kgCO₂e/m²)", result["net_intensity"]
        material_headers = [
            "Selected", "Category", "Material", "Method", "Declared Unit", "Declared Quantity", "Quantity Unit",
            "EPD Emitted / Unit", "EPD Stored / Unit", "Emitted Total", "Stored Total", "Net Total",
            "Emitted / m²", "Status", "Formula / Error", "Source",
        ]
        for col, value in enumerate(material_headers, 1):
            ws.cell(10, col, value)
        style_header(ws, 10, 1, len(material_headers))

        for material, row in zip(assembly["materials"], result["rows"]):
            ws.append([
                ("Yes" if material_is_selected(material) else "No"), row["category"], row["name"], row["method"], row["unit"],
                row["quantity"] if row["valid"] else None, row["quantity_unit"], row.get("epd_emitted"),
                row.get("epd_stored"), row["emitted"] if row["valid"] else None,
                row["stored"] if row["valid"] else None, row["net"] if row["valid"] else None,
                row["emitted_intensity"] if row["valid"] else None, row["status"],
                row["formula_note"] if row["valid"] else row["error"], row.get("source", ""),
            ])

        ws.freeze_panes = "A11"
        for col in range(1, len(material_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["C"].width = 42
        ws.column_dimensions["O"].width = 42
        ws.column_dimensions["P"].width = 38

    library = wb.create_sheet("Material Library")
    library_headers = ["Category", "Material", "Emitted A1–A3", "Stored Carbon", "Declared Unit", "Method", "Source", "Source Declared Unit"]
    library.append(library_headers)
    style_header(library, 1, 1, len(library_headers))
    for item in all_library_materials():
        library.append([item.get("category"), item.get("name"), item.get("emitted"), item.get("stored"), item.get("unit"), item.get("method"), item.get("source"), item.get("source_declared_unit")])
    library.freeze_panes = "A2"
    for col in range(1, len(library_headers) + 1):
        library.column_dimensions[get_column_letter(col)].width = 22
    library.column_dimensions["B"].width = 55
    library.column_dimensions["G"].width = 42
    library.column_dimensions["H"].width = 38

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def pdf_safe_text(value: Any) -> str:
    """Convert app text to a form supported reliably by ReportLab's built-in fonts."""
    text = str(value if value is not None else "")
    replacements = {
        "CO₂": "CO2",
        "₂": "2",
        "²": "2",
        "³": "3",
        "–": "-",
        "—": "-",
        "×": "x",
        "÷": "/",
        "•": "-",
        "≤": "<=",
        "≥": ">=",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = unicodedata.normalize("NFKD", text)
    return text.encode("latin-1", "replace").decode("latin-1")


def _chart_bounds(series: list[tuple[str, list[float], str]], stacked: bool) -> tuple[float, float]:
    if not series or not series[0][1]:
        return 0.0, 1.0

    count = len(series[0][1])
    if stacked:
        positive = []
        negative = []
        for index in range(count):
            positive.append(sum(max(0.0, values[index]) for _, values, _ in series))
            negative.append(sum(min(0.0, values[index]) for _, values, _ in series))
        minimum = min([0.0] + negative)
        maximum = max([0.0] + positive)
    else:
        values = [value for _, items, _ in series for value in items]
        minimum = min([0.0] + values)
        maximum = max([0.0] + values)

    if math.isclose(minimum, maximum):
        maximum = minimum + 1.0
    padding = (maximum - minimum) * 0.12
    return minimum - (padding if minimum < 0 else 0), maximum + (padding if maximum > 0 else 1.0)


def report_bar_chart(
    category_names: list[str],
    series: list[tuple[str, list[float], str]],
    y_axis_label: str,
    *,
    stacked: bool = False,
) -> Drawing:
    """Create a dependency-light vector bar chart for the downloadable PDF report."""
    drawing_width = 710
    drawing_height = 285
    drawing = Drawing(drawing_width, drawing_height)

    left = 62
    right = 18
    bottom = 66
    top = 42
    chart_width = drawing_width - left - right
    chart_height = drawing_height - bottom - top
    minimum, maximum = _chart_bounds(series, stacked)

    def y_position(value: float) -> float:
        return bottom + (value - minimum) / (maximum - minimum) * chart_height

    grid_color = rl_colors.HexColor(COLORS["line"])
    ink_color = rl_colors.HexColor(COLORS["ink"])
    muted_color = rl_colors.HexColor(COLORS["muted"])

    tick_count = 5
    for tick in range(tick_count + 1):
        value = minimum + (maximum - minimum) * tick / tick_count
        y = y_position(value)
        drawing.add(Line(left, y, left + chart_width, y, strokeColor=grid_color, strokeWidth=0.5))
        drawing.add(String(left - 7, y - 3, f"{value:,.1f}", textAnchor="end", fontName="Helvetica", fontSize=7, fillColor=muted_color))

    baseline = y_position(0.0)
    drawing.add(Line(left, baseline, left + chart_width, baseline, strokeColor=rl_colors.HexColor("#94A3B8"), strokeWidth=1.0))
    drawing.add(Line(left, bottom, left, bottom + chart_height, strokeColor=ink_color, strokeWidth=0.8))

    category_count = max(1, len(category_names))
    group_width = chart_width / category_count

    if stacked:
        bar_width = min(54.0, group_width * 0.52)
        for category_index in range(len(category_names)):
            x = left + category_index * group_width + (group_width - bar_width) / 2
            positive_total = 0.0
            negative_total = 0.0
            for _, values, hex_color in series:
                value = float(values[category_index])
                if value >= 0:
                    start = positive_total
                    positive_total += value
                    end = positive_total
                else:
                    start = negative_total
                    negative_total += value
                    end = negative_total
                y1 = y_position(start)
                y2 = y_position(end)
                drawing.add(Rect(x, min(y1, y2), bar_width, max(0.8, abs(y2 - y1)), fillColor=rl_colors.HexColor(hex_color), strokeColor=None))
    else:
        series_count = max(1, len(series))
        available_width = group_width * 0.72
        gap = min(4.0, available_width * 0.06)
        bar_width = max(2.0, min(30.0, (available_width - gap * (series_count - 1)) / series_count))
        total_width = bar_width * series_count + gap * (series_count - 1)
        for category_index in range(len(category_names)):
            start_x = left + category_index * group_width + (group_width - total_width) / 2
            for series_index, (_, values, hex_color) in enumerate(series):
                value = float(values[category_index])
                value_y = y_position(value)
                x = start_x + series_index * (bar_width + gap)
                drawing.add(Rect(x, min(baseline, value_y), bar_width, max(0.8, abs(value_y - baseline)), fillColor=rl_colors.HexColor(hex_color), strokeColor=None))

    for category_index, category_name in enumerate(category_names):
        x = left + (category_index + 0.5) * group_width
        label = pdf_safe_text(category_name)
        if len(label) > 20:
            label = label[:18] + "..."
        drawing.add(String(x + 2, bottom - 8, label, textAnchor="end", fontName="Helvetica", fontSize=7, fillColor=ink_color, angle=32))

    legend_x = left
    legend_y = drawing_height - 17
    for label, _, hex_color in series:
        safe_label = pdf_safe_text(label)
        item_width = 18 + min(155, max(45, len(safe_label) * 5.2))
        if legend_x + item_width > drawing_width - right:
            legend_x = left
            legend_y -= 15
        drawing.add(Rect(legend_x, legend_y - 7, 9, 9, fillColor=rl_colors.HexColor(hex_color), strokeColor=None))
        drawing.add(String(legend_x + 13, legend_y - 5, safe_label, fontName="Helvetica", fontSize=7.5, fillColor=ink_color))
        legend_x += item_width

    drawing.add(String(13, bottom + chart_height / 2, pdf_safe_text(y_axis_label), fontName="Helvetica-Bold", fontSize=8, fillColor=ink_color, angle=90, textAnchor="middle"))
    return drawing


def comparison_report_bytes() -> bytes:
    """Build a PDF containing comparison results, all comparison diagrams, and assembly details."""
    stream = io.BytesIO()
    generated_at = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(letter),
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.42 * inch,
        bottomMargin=0.42 * inch,
        title="Wall Assembly Embodied Carbon Report",
        author="Wall Assembly Embodied Carbon Calculator",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=rl_colors.HexColor(COLORS["blue"]),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        textColor=rl_colors.HexColor(COLORS["muted"]),
        spaceAfter=14,
    )
    heading_style = ParagraphStyle(
        "ReportHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        textColor=rl_colors.HexColor(COLORS["blue"]),
        spaceBefore=6,
        spaceAfter=8,
    )
    chart_heading_style = ParagraphStyle(
        "ChartHeading",
        parent=heading_style,
        alignment=TA_CENTER,
        spaceBefore=2,
        spaceAfter=2,
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11.5,
        textColor=rl_colors.HexColor(COLORS["ink"]),
    )

    comparison_rows: list[dict[str, Any]] = []
    assembly_results: list[tuple[dict[str, Any], dict[str, Any]]] = []
    detail_rows: list[dict[str, Any]] = []

    for assembly in st.session_state.assemblies:
        result = calculate_assembly(assembly)
        assembly_results.append((assembly, result))
        comparison_rows.append({
            "Assembly": assembly["name"],
            "Width (ft)": assembly["width_ft"],
            "Height (ft)": assembly["height_ft"],
            "Area (ft2)": result["area_ft2"],
            "Emitted kgCO2e": result["emitted"],
            "Stored kgCO2e": result["stored"],
            "Net kgCO2e": result["net"],
            "Emitted kgCO2e/m2": result["emitted_intensity"],
            "Net kgCO2e/m2": result["net_intensity"],
            "Materials ready": f'{result["ready_count"]}/{result["selected_count"]}',
            "Status": result["status"],
        })
        for row in result["rows"]:
            if row["valid"]:
                detail_rows.append({
                    "Assembly": assembly["name"],
                    "Category": row["category"],
                    "Emitted kgCO2e": row["emitted"],
                })

    comparison = pd.DataFrame(comparison_rows)
    started = comparison[comparison["Materials ready"] != "0/0"].copy() if not comparison.empty else comparison

    story: list[Any] = [
        Paragraph("Wall Assembly Embodied Carbon Report", title_style),
        Paragraph(
            pdf_safe_text(
                f"A1-A3 comparison generated {generated_at}. Emitted carbon and stored carbon are shown separately; net carbon equals emitted plus stored carbon."
            ),
            subtitle_style,
        ),
    ]

    if started.empty:
        story.append(Paragraph("No wall options currently contain selected materials. Add materials to generate comparison results and diagrams.", body_style))
    else:
        lowest_net = started.loc[started["Net kgCO2e/m2"].idxmin()]
        lowest_emitted = started.loc[started["Emitted kgCO2e/m2"].idxmin()]
        key_metrics = [
            ["Lowest net intensity", "Lowest emitted intensity", "Options started"],
            [
                f'{pdf_safe_text(lowest_net["Assembly"])}\n{lowest_net["Net kgCO2e/m2"]:,.3f} kgCO2e/m2',
                f'{pdf_safe_text(lowest_emitted["Assembly"])}\n{lowest_emitted["Emitted kgCO2e/m2"]:,.3f} kgCO2e/m2',
                str(len(started)),
            ],
        ]
        metric_table = Table(key_metrics, colWidths=[2.35 * inch] * 3, rowHeights=[0.26 * inch, 0.48 * inch])
        metric_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor(COLORS["soft_blue"])),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor(COLORS["blue"])),
            ("TEXTCOLOR", (0, 1), (-1, 1), rl_colors.HexColor(COLORS["ink"])),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, 1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor(COLORS["line"])),
            ("BOX", (0, 0), (-1, -1), 0.8, rl_colors.HexColor(COLORS["line"])),
        ]))
        story.extend([metric_table, Spacer(1, 12), Paragraph("Assembly comparison", heading_style)])

        summary_header = [
            "Assembly", "Width", "Height", "Area", "Ready", "Emitted", "Stored", "Net", "Emitted / m2", "Net / m2", "Status"
        ]
        summary_data: list[list[Any]] = [summary_header]
        for _, row in comparison.iterrows():
            summary_data.append([
                pdf_safe_text(row["Assembly"]),
                f'{row["Width (ft)"]:,.1f} ft',
                f'{row["Height (ft)"]:,.1f} ft',
                f'{row["Area (ft2)"]:,.1f} ft2',
                row["Materials ready"],
                f'{row["Emitted kgCO2e"]:,.2f}',
                f'{row["Stored kgCO2e"]:,.2f}',
                f'{row["Net kgCO2e"]:,.2f}',
                f'{row["Emitted kgCO2e/m2"]:,.3f}',
                f'{row["Net kgCO2e/m2"]:,.3f}',
                row["Status"],
            ])

        summary_table = Table(
            summary_data,
            repeatRows=1,
            colWidths=[1.38 * inch, 0.52 * inch, 0.52 * inch, 0.65 * inch, 0.5 * inch, 0.68 * inch, 0.68 * inch, 0.68 * inch, 0.78 * inch, 0.72 * inch, 0.64 * inch],
        )
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor(COLORS["blue"])),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.8),
            ("ALIGN", (1, 1), (-2, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
            ("GRID", (0, 0), (-1, -1), 0.35, rl_colors.HexColor(COLORS["line"])),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(summary_table)

        category_names = [pdf_safe_text(value) for value in started["Assembly"].tolist()]

        story.extend([
            PageBreak(),
            KeepTogether([
                Paragraph("Intensity comparison", chart_heading_style),
                report_bar_chart(
                    category_names,
                    [
                        ("Emitted kgCO2e/m2", started["Emitted kgCO2e/m2"].astype(float).tolist(), COLORS["purple"]),
                        ("Net kgCO2e/m2", started["Net kgCO2e/m2"].astype(float).tolist(), COLORS["blue"]),
                    ],
                    "kgCO2e/m2",
                ),
            ]),
            PageBreak(),
            KeepTogether([
                Paragraph("Total carbon", chart_heading_style),
                report_bar_chart(
                    category_names,
                    [
                        ("Emitted kgCO2e", started["Emitted kgCO2e"].astype(float).tolist(), COLORS["purple"]),
                        ("Stored kgCO2e", started["Stored kgCO2e"].astype(float).tolist(), COLORS["green"]),
                        ("Net kgCO2e", started["Net kgCO2e"].astype(float).tolist(), COLORS["blue"]),
                    ],
                    "kgCO2e",
                ),
            ]),
        ])

        if detail_rows:
            detail = pd.DataFrame(detail_rows)
            by_category = detail.groupby(["Assembly", "Category"], as_index=False)["Emitted kgCO2e"].sum()
            category_series: list[tuple[str, list[float], str]] = []
            for category in STARTING_CATEGORIES:
                values = []
                for assembly_name in started["Assembly"].tolist():
                    match = by_category[(by_category["Assembly"] == assembly_name) & (by_category["Category"] == category)]
                    values.append(float(match["Emitted kgCO2e"].sum()) if not match.empty else 0.0)
                if any(not math.isclose(value, 0.0) for value in values):
                    category_series.append((category, values, CATEGORY_COLORS.get(category, COLORS["muted"])))

            if category_series:
                story.extend([
                    PageBreak(),
                    KeepTogether([
                        Paragraph("Material breakdown by category", chart_heading_style),
                        report_bar_chart(category_names, category_series, "Emitted kgCO2e", stacked=True),
                        Spacer(1, 8),
                        Paragraph("Each bar shows emitted A1-A3 carbon, stacked by material category. Exact material-level values are included in the assembly detail pages.", body_style),
                    ]),
                ])

    for assembly, result in assembly_results:
        story.extend([
            PageBreak(),
            Paragraph(pdf_safe_text(assembly["name"]), heading_style),
        ])
        assembly_metrics = [
            ["Wall size", "Area", "Emitted A1-A3", "Stored carbon", "Net carbon", "Net intensity"],
            [
                f'{assembly["width_ft"]:,.1f} x {assembly["height_ft"]:,.1f} ft',
                f'{result["area_ft2"]:,.1f} ft2',
                f'{result["emitted"]:,.2f} kgCO2e',
                f'{result["stored"]:,.2f} kgCO2e',
                f'{result["net"]:,.2f} kgCO2e',
                f'{result["net_intensity"]:,.3f} kgCO2e/m2',
            ],
        ]
        assembly_metric_table = Table(assembly_metrics, colWidths=[1.17 * inch] * 6)
        assembly_metric_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor(COLORS["soft_blue"])),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor(COLORS["blue"])),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor(COLORS["line"])),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([assembly_metric_table, Spacer(1, 10)])

        material_data: list[list[Any]] = [[
            "Category", "Material", "Declared unit", "Quantity", "Emitted", "Stored", "Net", "Status / formula"
        ]]
        for row in result["rows"]:
            if row["status"] == "Not selected":
                continue
            material_data.append([
                pdf_safe_text(row["category"]),
                pdf_safe_text(row["name"]),
                pdf_safe_text(row["unit"]),
                f'{row["quantity"]:,.3f} {pdf_safe_text(row["quantity_unit"])}' if row["valid"] else "-",
                f'{row["emitted"]:,.2f}' if row["valid"] else "-",
                f'{row["stored"]:,.2f}' if row["valid"] else "-",
                f'{row["net"]:,.2f}' if row["valid"] else "-",
                pdf_safe_text(row["formula_note"] if row["valid"] else row["error"] or row["status"]),
            ])

        if len(material_data) == 1:
            story.append(Paragraph("No materials are selected for this wall option.", body_style))
        else:
            material_table = Table(
                material_data,
                repeatRows=1,
                colWidths=[1.05 * inch, 2.1 * inch, 0.82 * inch, 0.95 * inch, 0.65 * inch, 0.65 * inch, 0.65 * inch, 1.45 * inch],
            )
            material_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor(COLORS["blue"])),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("ALIGN", (3, 1), (6, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
                ("GRID", (0, 0), (-1, -1), 0.35, rl_colors.HexColor(COLORS["line"])),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(material_table)

    def add_footer(canvas, document) -> None:
        canvas.saveState()
        canvas.setStrokeColor(rl_colors.HexColor(COLORS["line"]))
        canvas.line(doc.leftMargin, 20, landscape(letter)[0] - doc.rightMargin, 20)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(rl_colors.HexColor(COLORS["muted"]))
        canvas.drawString(doc.leftMargin, 9, "Wall Assembly Embodied Carbon Calculator - A1-A3 report")
        canvas.drawRightString(landscape(letter)[0] - doc.rightMargin, 9, f"Page {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    return stream.getvalue()

# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
    st.markdown('<div class="sidebar-brand">Wall Assembly Embodied Carbon Calculator</div>', unsafe_allow_html=True)
    page_options = ["Build Assemblies", "Compare Assemblies", "Material Library", "Definitions", "Method Guide"]
    page = st.radio(
        "Navigation",
        page_options,
        index=select_index(page_options, st.session_state.page),
        label_visibility="collapsed",
    )
    st.session_state.page = page

    st.divider()
    st.markdown("### Export")
    st.download_button(
        "Download results Excel",
        data=workbook_bytes(),
        file_name="wall_assembly_carbon_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    st.download_button(
        "Download comparison report PDF",
        data=comparison_report_bytes(),
        file_name="wall_assembly_carbon_report.pdf",
        mime="application/pdf",
        width="stretch",
    )

    st.divider()
    st.caption("A1–A3 emitted carbon and stored carbon are shown separately. Net carbon = emitted + stored.")

# =============================================================================
# PAGE: BUILD ASSEMBLIES
# =============================================================================

if page == "Build Assemblies":
    st.markdown(
        """
        <div class="hero">
          <h1>Wall Assembly Embodied Carbon Calculator</h1>
        </div>
        """,
        unsafe_allow_html=True,
    )

    assembly_labels = [f"{index + 1}. {assembly['name']}" for index, assembly in enumerate(st.session_state.assemblies)]
    active = get_active_assembly()
    active_index = next(index for index, assembly in enumerate(st.session_state.assemblies) if assembly["id"] == active["id"])
    top1, top2, top3, top4 = st.columns([4.6, 1.25, 1.35, 1.15])
    with top1:
        selected_label = st.selectbox("Wall option", assembly_labels, index=active_index)
        selected_id = st.session_state.assemblies[assembly_labels.index(selected_label)]["id"]
        if selected_id != st.session_state.active_assembly_id:
            st.session_state.active_assembly_id = selected_id
            st.rerun()
    with top2:
        st.write("")
        if st.button("＋ New", width="stretch", type="primary"):
            created = new_assembly(len(st.session_state.assemblies) + 1)
            st.session_state.assemblies.append(created)
            st.session_state.active_assembly_id = created["id"]
            st.rerun()
    with top3:
        st.write("")
        if st.button("⧉ Duplicate", width="stretch"):
            duplicated = copy.deepcopy(active)
            duplicated["id"] = uuid.uuid4().hex
            duplicated["name"] = f'{active["name"]} Copy'
            for material in duplicated["materials"]:
                material["id"] = uuid.uuid4().hex
            st.session_state.assemblies.append(normalize_assembly(duplicated))
            st.session_state.active_assembly_id = duplicated["id"]
            st.rerun()
    with top4:
        st.write("")
        if st.button("Delete", width="stretch", disabled=len(st.session_state.assemblies) == 1):
            st.session_state.assemblies = [item for item in st.session_state.assemblies if item["id"] != active["id"]]
            st.session_state.active_assembly_id = st.session_state.assemblies[0]["id"]
            st.rerun()

    active = normalize_assembly(get_active_assembly())
    with st.container(border=True):
        st.markdown('<div class="section-label">Wall dimensions</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns([2.2, 1, 1])
        with c1:
            active["name"] = st.text_input("Assembly name", value=active["name"], key=f'assembly_name_{active["id"]}')
        with c2:
            active["width_ft"] = st.number_input("Wall width (ft)", min_value=0.0, value=float(active["width_ft"]), step=1.0, key=f'width_{active["id"]}')
        with c3:
            active["height_ft"] = st.number_input("Wall height (ft)", min_value=0.0, value=float(active["height_ft"]), step=1.0, key=f'height_{active["id"]}')

    summary_placeholder = st.empty()

    st.markdown("### Wall layers")

    delete_id = None
    for category in STARTING_CATEGORIES:
        category_materials = [item for item in active["materials"] if item["category"] == category]
        with st.container(border=True):
            title_col, add_col = st.columns([5, 1.25])
            with title_col:
                color = CATEGORY_COLORS[category]
                st.markdown(
                    f'<div class="category-heading"><span class="category-dot" style="background:{color}"></span>'
                    f'<span class="category-title">{category}</span></div>',
                    unsafe_allow_html=True,
                )
            with add_col:
                if st.button("＋ Add layer", key=f'add_{active["id"]}_{category}', width="stretch"):
                    active["materials"].append(new_material(category, is_base=False))
                    st.rerun()

            for material in category_materials:
                existing = calculate_material(active, material)
                selected = material_is_selected(material)
                title = existing["name"] if selected else "None selected"
                expanded = selected and not existing["valid"]

                with st.expander(title, expanded=expanded):
                    prefix = f'{active["id"]}_{material["id"]}'
                    selector_col, remove_col = st.columns([5, 1])
                    with selector_col:
                        choices = presets_for_category(category)
                        current = material.get("preset", "None")
                        if current not in choices:
                            current = "None"
                        material["preset"] = st.selectbox(
                            "Material / EPD",
                            choices,
                            index=select_index(choices, current),
                            key=f'{prefix}_preset',
                        )
                    with remove_col:
                        st.write("")
                        if not material.get("is_base", False) and st.button("Remove", key=f'{prefix}_remove', width="stretch"):
                            delete_id = material["id"]

                    if material["preset"] == "None":
                        continue

                    if material["preset"] == "Custom EPD":
                        st.markdown('<div class="section-label">Custom EPD</div>', unsafe_allow_html=True)
                        material["custom_name"] = st.text_input("Material name", value=material.get("custom_name", ""), key=f'{prefix}_custom_name')
                        c1, c2 = st.columns(2)
                        with c1:
                            material["custom_emitted"] = st.number_input("Emitted A1–A3 per declared unit", value=float(material.get("custom_emitted", 0.0)), step=0.1, key=f'{prefix}_custom_emitted')
                        with c2:
                            material["custom_stored"] = st.number_input("Stored carbon per declared unit", value=float(material.get("custom_stored", 0.0)), step=0.1, key=f'{prefix}_custom_stored', help="Use a negative value for stored or sequestered carbon.")
                        material["custom_unit"] = st.selectbox(
                            "Declared unit",
                            DECLARED_UNITS,
                            index=select_index(DECLARED_UNITS, material.get("custom_unit", DECLARED_UNITS[0])),
                            key=f'{prefix}_custom_unit',
                        )
                        st.caption(
                            f'Formula selected automatically: '
                            f'{method_from_declared_unit(material["custom_unit"])}'
                        )
                        if st.button("Add to custom material library", key=f'{prefix}_save_custom_library'):
                            custom_name = material.get("custom_name", "").strip()
                            if not custom_name:
                                st.error("Enter a material name before adding it to the custom library.")
                            elif material_name_exists(custom_name):
                                st.error("A preset or custom material with that name already exists.")
                            else:
                                st.session_state.custom_library.append({
                                    "id": uuid.uuid4().hex,
                                    "category": category,
                                    "name": custom_name,
                                    "emitted": float(material.get("custom_emitted", 0.0)),
                                    "stored": float(material.get("custom_stored", 0.0)),
                                    "unit": material.get("custom_unit", DECLARED_UNITS[0]),
                                    "method": method_from_declared_unit(material.get("custom_unit", DECLARED_UNITS[0])),
                                    "source": "Saved from assembly builder",
                                    "source_declared_unit": "",
                                })
                                mark_custom_library_changed()
                                st.success("Custom material added. You can edit or remove it from Material Library.")
                                st.rerun()
                        epd = effective_epd(material)
                    else:
                        epd = effective_epd(material)
                        if epd:
                            verification = "Verified preset" if "Verified" in epd.get("source", "") else "Preset material"
                            st.markdown(
                                f'<div class="epd-card"><strong>{verification}</strong><br>'
                                f'<span class="small-muted">Emitted: {epd["emitted"]:,.6g} {epd["unit"]} &nbsp;•&nbsp; '
                                f'Stored: {epd["stored"]:,.6g} {epd["unit"]}</span></div>',
                                unsafe_allow_html=True,
                            )

                    if epd:
                        render_method_inputs(material, epd["method"], prefix)
                        row = calculate_material(active, material)
                        if row["valid"]:
                            st.markdown('<div class="section-label">Result</div>', unsafe_allow_html=True)
                            r1, r2, r3, r4 = st.columns(4)
                            r1.metric("Quantity", f'{row["quantity"]:,.3f} {row["quantity_unit"]}')
                            r2.metric("Emitted", f'{row["emitted"]:,.2f} kgCO₂e')
                            r3.metric("Stored", f'{row["stored"]:,.2f} kgCO₂e')
                            r4.metric("Net", f'{row["net"]:,.2f} kgCO₂e')
                            st.markdown(f'<div class="formula-box">{row["formula_note"]}</div>', unsafe_allow_html=True)
                        else:
                            st.error(row["error"] or row["status"])

    if delete_id:
        active["materials"] = [item for item in active["materials"] if item["id"] != delete_id]
        st.rerun()

    result = calculate_assembly(active)
    with summary_placeholder.container():
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Wall area", f'{result["area_ft2"]:,.1f} ft²')
        c2.metric("Emitted A1–A3", f'{result["emitted"]:,.1f} kgCO₂e')
        c3.metric("Stored carbon", f'{result["stored"]:,.1f} kgCO₂e')
        c4.metric("Net carbon", f'{result["net"]:,.1f} kgCO₂e')
        c5.metric("Net intensity", f'{result["net_intensity"]:,.3f} kgCO₂e/m²')
    st.caption(f'{result["ready_count"]} of {result["selected_count"]} selected layers are ready.')

    ready_rows = [row for row in result["rows"] if row["valid"]]
    if ready_rows:
        st.markdown("### Material contribution")
        chart_df = pd.DataFrame([
            {"Material": row["name"], "Category": row["category"], "Emitted kgCO₂e": row["emitted"]}
            for row in ready_rows
        ]).sort_values("Emitted kgCO₂e", ascending=True)
        fig = px.bar(
            chart_df,
            x="Emitted kgCO₂e",
            y="Material",
            color="Category",
            orientation="h",
            color_discrete_map=CATEGORY_COLORS,
            height=max(340, 42 * len(chart_df)),
        )
        fig.update_layout(
            font=dict(family="Roboto, Arial, sans-serif", color=COLORS["ink"]),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            legend_title_text="",
            margin=dict(l=10, r=10, t=15, b=10),
            xaxis_title="kgCO₂e",
            yaxis_title="",
        )
        fig.update_xaxes(gridcolor=COLORS["line"])
        st.plotly_chart(fig, width="stretch")

# =============================================================================
# PAGE: COMPARE ASSEMBLIES
# =============================================================================

elif page == "Compare Assemblies":
    st.markdown(
        """
        <div class="hero">
          <h1>Compare Wall Assemblies</h1>
          <p>Review total carbon, carbon intensity, and material contributions across every named wall option.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    comparison_rows = []
    detail_rows = []
    for assembly in st.session_state.assemblies:
        result = calculate_assembly(assembly)
        comparison_rows.append({
            "Assembly": assembly["name"],
            "Width (ft)": assembly["width_ft"],
            "Height (ft)": assembly["height_ft"],
            "Area (ft²)": result["area_ft2"],
            "Emitted kgCO₂e": result["emitted"],
            "Stored kgCO₂e": result["stored"],
            "Net kgCO₂e": result["net"],
            "Emitted kgCO₂e/m²": result["emitted_intensity"],
            "Net kgCO₂e/m²": result["net_intensity"],
            "Materials ready": f'{result["ready_count"]}/{result["selected_count"]}',
            "Status": result["status"],
        })
        for row in result["rows"]:
            if row["valid"]:
                detail_rows.append({
                    "Assembly": assembly["name"], "Material": row["name"], "Category": row["category"],
                    "Emitted kgCO₂e": row["emitted"], "Stored kgCO₂e": row["stored"], "Net kgCO₂e": row["net"],
                })

    comparison = pd.DataFrame(comparison_rows)
    started = comparison[comparison["Materials ready"] != "0/0"].copy()

    if started.empty:
        st.info("Add materials to at least one wall option to see comparisons.")
    else:
        lowest_net = started.loc[started["Net kgCO₂e/m²"].idxmin()]
        lowest_emitted = started.loc[started["Emitted kgCO₂e/m²"].idxmin()]
        c1, c2, c3 = st.columns(3)
        c1.metric("Lowest net intensity", lowest_net["Assembly"], f'{lowest_net["Net kgCO₂e/m²"]:,.3f} kgCO₂e/m²')
        c2.metric("Lowest emitted intensity", lowest_emitted["Assembly"], f'{lowest_emitted["Emitted kgCO₂e/m²"]:,.3f} kgCO₂e/m²')
        c3.metric("Options started", f'{len(started)}')

        tab1, tab2, tab3 = st.tabs(["Intensity comparison", "Total carbon", "Material breakdown"])
        with tab1:
            intensity_measure_order = ["Emitted kgCO₂e/m²", "Net kgCO₂e/m²"]
            intensity_long = started.melt(
                id_vars=["Assembly"],
                value_vars=intensity_measure_order,
                var_name="Measure", value_name="kgCO₂e/m²",
            )
            fig = px.bar(
                intensity_long,
                x="Assembly",
                y="kgCO₂e/m²",
                color="Measure",
                text="kgCO₂e/m²",
                barmode="group",
                category_orders={
                    "Assembly": started["Assembly"].tolist(),
                    "Measure": intensity_measure_order,
                },
                color_discrete_sequence=[COLORS["purple"], COLORS["blue"]],
                height=520,
            )
            fig.update_traces(
                texttemplate="%{y:,.3f}",
                textposition="outside",
                textfont=dict(size=16),
                cliponaxis=False,
            )
            fig.update_layout(
                title=dict(
                    text="Intensity comparison",
                    font=dict(size=22, color=COLORS["blue"]),
                    x=0.01,
                    xanchor="left",
                ),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(family="Roboto, Arial, sans-serif", size=14, color=COLORS["ink"]),
                legend_title_text="",
                legend=dict(font=dict(size=16), traceorder="normal"),
                xaxis_title="",
                margin=dict(t=80, r=30, b=45, l=55),
                uniformtext_minsize=14,
                uniformtext_mode="show",
            )
            fig.update_yaxes(gridcolor=COLORS["line"])
            st.plotly_chart(fig, width="stretch")

        with tab2:
            total_measure_order = ["Emitted kgCO₂e", "Stored kgCO₂e", "Net kgCO₂e"]
            total_long = started.melt(
                id_vars=["Assembly"],
                value_vars=total_measure_order,
                var_name="Measure", value_name="kgCO₂e",
            )
            fig = px.bar(
                total_long,
                x="Assembly",
                y="kgCO₂e",
                color="Measure",
                text="kgCO₂e",
                barmode="group",
                category_orders={
                    "Assembly": started["Assembly"].tolist(),
                    "Measure": total_measure_order,
                },
                color_discrete_sequence=[COLORS["purple"], COLORS["green"], COLORS["blue"]],
                height=520,
            )
            fig.update_traces(
                texttemplate="%{y:,.2f}",
                textposition="outside",
                textfont=dict(size=16),
                cliponaxis=False,
            )
            fig.update_layout(
                title=dict(
                    text="Total carbon",
                    font=dict(size=22, color=COLORS["blue"]),
                    x=0.01,
                    xanchor="left",
                ),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(family="Roboto, Arial, sans-serif", size=14, color=COLORS["ink"]),
                legend_title_text="",
                legend=dict(font=dict(size=16), traceorder="normal"),
                xaxis_title="",
                margin=dict(t=80, r=30, b=45, l=55),
                uniformtext_minsize=14,
                uniformtext_mode="show",
            )
            fig.update_yaxes(gridcolor=COLORS["line"], zerolinecolor="#94A3B8")
            st.plotly_chart(fig, width="stretch")

        with tab3:
            if detail_rows:
                detail = pd.DataFrame(detail_rows)
                by_category = detail.groupby(["Assembly", "Category"], as_index=False)["Emitted kgCO₂e"].sum()

                # Keep the stack in the wall-layer order. Plotly stacks traces from
                # bottom to top, so the legend is reversed to read in the same
                # top-to-bottom order as the visible stacked bar.
                material_breakdown_order = [
                    category for category in STARTING_CATEGORIES
                    if category in set(by_category["Category"])
                ]
                by_category["Category"] = pd.Categorical(
                    by_category["Category"],
                    categories=material_breakdown_order,
                    ordered=True,
                )
                by_category = by_category.sort_values(["Assembly", "Category"])

                # Only place a number inside a stack segment when the segment is
                # tall enough to hold it. Smaller contributions remain available in
                # the hover tooltip instead of overlapping neighbouring labels.
                assembly_totals = by_category.groupby("Assembly")["Emitted kgCO₂e"].sum()
                largest_total = float(assembly_totals.max()) if not assembly_totals.empty else 0.0
                label_threshold = max(20.0, largest_total * 0.035)
                by_category["Chart label"] = by_category["Emitted kgCO₂e"].map(
                    lambda value: f"{value:,.2f}" if abs(float(value)) >= label_threshold else ""
                )

                fig = px.bar(
                    by_category,
                    x="Assembly",
                    y="Emitted kgCO₂e",
                    color="Category",
                    text="Chart label",
                    barmode="stack",
                    category_orders={
                        "Assembly": started["Assembly"].tolist(),
                        "Category": material_breakdown_order,
                    },
                    color_discrete_map=CATEGORY_COLORS,
                    height=570,
                )
                fig.update_traces(
                    texttemplate="%{text}",
                    textposition="inside",
                    textfont=dict(size=13),
                    insidetextanchor="middle",
                    cliponaxis=True,
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "Category: %{fullData.name}<br>"
                        "Emitted: %{y:,.2f} kgCO₂e<extra></extra>"
                    ),
                )
                fig.update_layout(
                    title=dict(
                        text="Material breakdown",
                        font=dict(size=22, color=COLORS["blue"]),
                        x=0.01,
                        xanchor="left",
                    ),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Roboto, Arial, sans-serif", size=14, color=COLORS["ink"]),
                    legend_title_text="",
                    legend=dict(
                        font=dict(size=16),
                        traceorder="reversed",
                        x=0.99,
                        xanchor="right",
                        y=1.0,
                        yanchor="top",
                        bgcolor="rgba(255,255,255,0.88)",
                        bordercolor=COLORS["line"],
                        borderwidth=1,
                    ),
                    xaxis=dict(
                        title="",
                        domain=[0.0, 0.76],
                    ),
                    margin=dict(t=80, r=25, b=45, l=60),
                    uniformtext_minsize=10,
                    uniformtext_mode="hide",
                )
                fig.update_yaxes(gridcolor=COLORS["line"])
                st.plotly_chart(fig, width="stretch")

        st.markdown("### Comparison table")
        format_columns = {
            "Area (ft²)": "{:.1f}", "Emitted kgCO₂e": "{:.2f}", "Stored kgCO₂e": "{:.2f}",
            "Net kgCO₂e": "{:.2f}", "Emitted kgCO₂e/m²": "{:.3f}", "Net kgCO₂e/m²": "{:.3f}",
        }
        st.dataframe(comparison.style.format(format_columns), width="stretch", hide_index=True)

# =============================================================================
# PAGE: MATERIAL LIBRARY
# =============================================================================

elif page == "Material Library":
    st.markdown(
        """
        <div class="hero">
          <h1>Material Library</h1>
          <p>Browse preset EPD factors and manage your reusable custom materials.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        link_text, link_button = st.columns([4, 1.25])
        with link_text:
            st.markdown("**Looking for more materials?**")
            st.caption("Browse additional EPDs and material data through 2050 Materials.")
        with link_button:
            st.write("")
            st.link_button(
                "Open 2050 Materials",
                "https://app.2050-materials.com/",
                width="stretch",
            )

    custom_header, custom_add = st.columns([5, 1.25])
    with custom_header:
        st.markdown("### Custom materials")
        st.caption("Custom materials are saved in this browser and remain available after refreshing or returning to the app.")
    with custom_add:
        st.write("")
        if st.button("＋ Add custom", width="stretch", type="primary"):
            st.session_state.show_add_custom = not st.session_state.show_add_custom
            st.session_state.editing_custom_id = None
            st.rerun()

    if st.session_state.show_add_custom:
        with st.container(border=True):
            st.markdown("#### Add custom material")
            with st.form("add_custom_material_form", clear_on_submit=False):
                c1, c2 = st.columns([1, 2])
                add_category = c1.selectbox("Category", category_list(), key="add_custom_category")
                add_name = c2.text_input("Material name", key="add_custom_name")
                c1, c2 = st.columns(2)
                add_emitted = c1.number_input("Emitted A1–A3 per declared unit", value=0.0, step=0.1, key="add_custom_emitted")
                add_stored = c2.number_input("Stored carbon per declared unit", value=0.0, step=0.1, key="add_custom_stored")
                add_unit = st.selectbox("Declared unit", DECLARED_UNITS, key="add_custom_unit")
                st.caption(f"Formula selected automatically: {method_from_declared_unit(add_unit)}")
                add_source = st.text_input("Source / verification (optional)", key="add_custom_source")
                b1, b2 = st.columns(2)
                add_submitted = b1.form_submit_button("Add material", type="primary", width="stretch")
                add_cancelled = b2.form_submit_button("Cancel", width="stretch")

            if add_cancelled:
                st.session_state.show_add_custom = False
                st.rerun()
            if add_submitted:
                clean_name = add_name.strip()
                if not clean_name:
                    st.error("Enter a material name.")
                elif material_name_exists(clean_name):
                    st.error("A preset or custom material with that name already exists.")
                else:
                    st.session_state.custom_library.append({
                        "id": uuid.uuid4().hex,
                        "category": add_category,
                        "name": clean_name,
                        "emitted": float(add_emitted),
                        "stored": float(add_stored),
                        "unit": add_unit,
                        "method": method_from_declared_unit(add_unit),
                        "source": add_source.strip() or "Custom material",
                        "source_declared_unit": "",
                    })
                    st.session_state.show_add_custom = False
                    mark_custom_library_changed()
                    st.rerun()

    if not st.session_state.custom_library:
        st.info("No custom materials added yet.")
    else:
        for custom in list(st.session_state.custom_library):
            custom_id = custom["id"]
            with st.container(border=True):
                info_col, edit_col, remove_col = st.columns([5, 1, 1])
                with info_col:
                    st.markdown(f'**{custom["name"]}**')
                    st.caption(
                        f'{custom["category"]} · {custom["emitted"]:,.6g} emitted · '
                        f'{custom["stored"]:,.6g} stored · {custom["unit"]}'
                    )
                with edit_col:
                    if st.button("Edit", key=f"edit_custom_{custom_id}", width="stretch"):
                        st.session_state.editing_custom_id = custom_id
                        st.session_state.show_add_custom = False
                        st.rerun()
                with remove_col:
                    if st.button("Remove", key=f"remove_custom_{custom_id}", width="stretch"):
                        update_custom_material_references(custom["name"], None)
                        st.session_state.custom_library = [
                            item for item in st.session_state.custom_library if item.get("id") != custom_id
                        ]
                        if st.session_state.editing_custom_id == custom_id:
                            st.session_state.editing_custom_id = None
                        mark_custom_library_changed()
                        st.rerun()

                if st.session_state.editing_custom_id == custom_id:
                    st.divider()
                    with st.form(f"edit_custom_form_{custom_id}"):
                        c1, c2 = st.columns([1, 2])
                        edit_category = c1.selectbox(
                            "Category",
                            category_list(),
                            index=select_index(category_list(), custom["category"]),
                            key=f"edit_category_{custom_id}",
                        )
                        edit_name = c2.text_input("Material name", value=custom["name"], key=f"edit_name_{custom_id}")
                        c1, c2 = st.columns(2)
                        edit_emitted = c1.number_input(
                            "Emitted A1–A3 per declared unit",
                            value=float(custom["emitted"]),
                            step=0.1,
                            key=f"edit_emitted_{custom_id}",
                        )
                        edit_stored = c2.number_input(
                            "Stored carbon per declared unit",
                            value=float(custom["stored"]),
                            step=0.1,
                            key=f"edit_stored_{custom_id}",
                        )
                        edit_unit = st.selectbox(
                            "Declared unit",
                            DECLARED_UNITS,
                            index=select_index(DECLARED_UNITS, custom["unit"]),
                            key=f"edit_unit_{custom_id}",
                        )
                        st.caption(f"Formula selected automatically: {method_from_declared_unit(edit_unit)}")
                        edit_source = st.text_input(
                            "Source / verification (optional)",
                            value=custom.get("source", ""),
                            key=f"edit_source_{custom_id}",
                        )
                        b1, b2 = st.columns(2)
                        edit_saved = b1.form_submit_button("Save changes", type="primary", width="stretch")
                        edit_cancelled = b2.form_submit_button("Cancel", width="stretch")

                    if edit_cancelled:
                        st.session_state.editing_custom_id = None
                        st.rerun()
                    if edit_saved:
                        clean_name = edit_name.strip()
                        if not clean_name:
                            st.error("Enter a material name.")
                        elif material_name_exists(clean_name, exclude_custom_id=custom_id):
                            st.error("A preset or custom material with that name already exists.")
                        else:
                            old_name = custom["name"]
                            old_category = custom["category"]
                            custom.update({
                                "category": edit_category,
                                "name": clean_name,
                                "emitted": float(edit_emitted),
                                "stored": float(edit_stored),
                                "unit": edit_unit,
                                "method": method_from_declared_unit(edit_unit),
                                "source": edit_source.strip() or "Custom material",
                                "source_declared_unit": "",
                            })
                            if old_category == edit_category:
                                update_custom_material_references(old_name, clean_name)
                            else:
                                update_custom_material_references(old_name, None)
                            st.session_state.editing_custom_id = None
                            mark_custom_library_changed()
                            st.rerun()

    st.markdown("### Full material library")
    library_rows = []
    for item in PRESET_MATERIALS:
        library_rows.append({**item, "library_type": "Preset"})
    for item in st.session_state.custom_library:
        library_rows.append({**item, "library_type": "Custom"})
    library_df = pd.DataFrame(library_rows)

    f1, f2, f3 = st.columns([2, 1.2, 1.2])
    with f1:
        search = st.text_input("Search materials", placeholder="Manufacturer, product, EPD, category…")
    with f2:
        categories = ["All"] + category_list()
        category_filter = st.selectbox("Category", categories)
    with f3:
        library_type_filter = st.selectbox("Library type", ["All", "Preset", "Custom"])

    filtered = library_df.copy()
    if search:
        mask = filtered.astype(str).apply(lambda col: col.str.contains(search, case=False, na=False)).any(axis=1)
        filtered = filtered[mask]
    if category_filter != "All":
        filtered = filtered[filtered["category"] == category_filter]
    if library_type_filter != "All":
        filtered = filtered[filtered["library_type"] == library_type_filter]

    shown = filtered.rename(columns={
        "library_type": "Type",
        "category": "Category",
        "name": "Material",
        "emitted": "Emitted A1–A3 / unit",
        "stored": "Stored carbon / unit",
        "unit": "Declared unit",
        "method": "Automatic formula",
        "source": "Source / verification",
        "source_declared_unit": "Source declared unit",
    })
    display_columns = [
        "Type", "Category", "Material", "Emitted A1–A3 / unit", "Stored carbon / unit",
        "Declared unit", "Automatic formula", "Source / verification", "Source declared unit",
    ]
    shown = shown[[column for column in display_columns if column in shown.columns]]
    st.dataframe(shown, width="stretch", hide_index=True)
    st.download_button(
        "Download filtered library CSV",
        shown.to_csv(index=False).encode("utf-8"),
        "material_library.csv",
        "text/csv",
    )

# =============================================================================
# PAGE: DEFINITIONS
# =============================================================================

elif page == "Definitions":
    st.markdown(
        """
        <div class="hero">
          <h1>Definitions</h1>
          <p>Simple terms used throughout the calculator.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    definitions = [
        (
            "What is A1–A3?",
            "A1–A3 is the product stage of a material's life cycle: raw-material supply (A1), transport to manufacturing (A2), and manufacturing (A3).",
        ),
        (
            "Why does this calculator use only A1–A3?",
            "A1–A3 provides a consistent comparison of the upfront product impacts reported by EPDs. Later stages depend more on project location, construction, maintenance, replacement, demolition, and end-of-life assumptions. This is therefore an upfront product-stage comparison, not a full whole-life assessment.",
        ),
        (
            "What is the EPD declared unit, and where do I find it?",
            "The declared unit is the quantity of product used as the basis for the EPD results, such as 1 m², 1 m³, 1 kg, 1 metre, or 1 unit. Look near the beginning of the EPD under “Declared unit,” “Functional unit,” or “Product information.” It is also usually repeated above or beside the environmental-impact results table.",
        ),
        (
            "What is RSI, and how do I convert it to R-value?",
            "RSI is the metric measure of thermal resistance in m²·K/W. R-value is the imperial measure in h·ft²·°F/Btu. Convert using: R-value = RSI × 5.678. Convert back using: RSI = R-value × 0.1761.",
        ),
        (
            "What is stored or biogenic carbon?",
            "Stored (biogenic) carbon is carbon captured from the atmosphere via photosynthesis and temporarily retained within bio-based materials such as wood, hemp, or paper. When an EPD reports stored carbon, enter it as a negative value in this calculator.",
        ),
        (
            "What is net carbon?",
            "Net carbon = emitted A1–A3 carbon + stored carbon. A negative stored-carbon value reduces the net result.",
        ),
        (
            "What is carbon intensity?",
            "Carbon intensity is the emitted A1–A3 carbon divided by the wall area. It is shown in kgCO₂e/m² so walls of different sizes can be compared.",
        ),
        (
            "What is net intensity?",
            "Net intensity is net carbon divided by the wall area. It includes both emitted A1–A3 carbon and reported stored carbon, and is shown in kgCO₂e/m².",
        ),
    ]

    for title, definition in definitions:
        with st.container(border=True):
            st.markdown(f"### {title}")
            st.write(definition)

# =============================================================================
# PAGE: METHOD GUIDE
# =============================================================================

else:
    st.markdown(
        """
        <div class="hero">
          <h1>Calculation Method Guide</h1>
          <p>These formulas reproduce the connected Excel calculator without openings, waste percentages, or hidden helper sheets.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    guides = [
        ("Area", "kgCO₂e/m²", "GWP = EPD Emitted / m² × wall area in m²"),
        ("Volume Layer", "kgCO₂e/m³", "GWP = EPD Emitted / m³ ×  wall area in m² × thickness in meters × conversion inch to meters"),
        ("Mass - total kg or grammage", "kgCO₂e/kg", "Option 1: GWP = EPD Emitted /kg × total mass. Option 2: GWP = EPD Emitted /kg × wall area in m² × grammage in kg/m²"),
        ("Linear Members", "kgCO₂e/m", "GWP = EPD Emitted /m × member count × member length in meters"),
        ("Unit Grid", "kgCO₂e/unit", "GWP = EPD Emitted /m × member count × member length in meters"),
        ("Insulation R-value", "kgCO₂e/m² at R/RSI", "GWP = EPD Emitted /m² × required RSI ÷ EPD reference RSI"),
        ("Wood Stud Framing", "kgCO₂e/m³", "GWP = EPD Emitted / m³×  stud count × wall height × actual lumber cross-section"),
        ("Steel Stud Framing", "kgCO₂e/kg", "GWP = total stud length × approximate C-section steel area × steel density"),
    ]

    for method, unit, formula in guides:
        with st.container(border=True):
            c1, c2 = st.columns([1.2, 3])
            c1.markdown(f"### {method}")
            c1.caption(unit)
            c2.markdown(f'<div class="formula-box">{formula}</div>', unsafe_allow_html=True)

    st.warning("The steel-stud method is an approximation based on gauge thickness, stud depth, standard flange/lip assumptions, and steel density. Manufacturer weight-per-length data is preferred when available; use Mass by Linear Weight when you have it.")
    st.info("Stored carbon should normally be entered as a negative EPD value. Net carbon is calculated as emitted carbon + stored carbon.")
